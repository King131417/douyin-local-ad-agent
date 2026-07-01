"""
6月月报生成脚本
生成 Excel 文件，包含：
1. 总览
2. 每日趋势
3. 账户排行
4. 素材 TOP50
5. 投放单元 TOP30
"""
import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

DB_PATH = 'data/ad_data.db'
START_DATE = '2026-06-01'
END_DATE = '2026-06-28'
OUTPUT_DIR = 'outputs'

# ── 颜色 ──────────────────────────────────────────────────────────────────
DARK_BLUE = 'FF1a3c6b'
MID_BLUE  = 'FF2d6abf'
LIGHT_BLUE= 'FFdce8f5'
ORANGE    = 'FFe8732a'
GREEN     = 'FF27ae60'
RED       = 'FFc0392b'
GREY_BG   = 'FFF4F6F9'
WHITE     = 'FFFFFFFF'
YELLOW    = 'FFFFF3cd'

def style_header(cell, bg=DARK_BLUE, fg=WHITE, bold=True, size=10, center=True):
    cell.font = Font(name='微软雅黑', bold=bold, color=fg[2:], size=size)
    cell.fill = PatternFill('solid', fgColor=bg[2:])
    cell.alignment = Alignment(horizontal='center' if center else 'left',
                                vertical='center', wrap_text=True)

def style_cell(cell, bold=False, size=10, color=None, bg=None, center=False, wrap=False):
    cell.font = Font(name='微软雅黑', bold=bold, size=size,
                     color=color[2:] if color else '000000')
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg[2:])
    cell.alignment = Alignment(horizontal='center' if center else 'left',
                                vertical='center', wrap_text=wrap)

def thin_border():
    s = Side(border_style='thin', color='BBBBBB')
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell):
    ws.freeze_panes = cell

# ────────────────────────────────────────────────────────────────────────────
def build(conn):
    wb = Workbook()

    # ─── Sheet 1: 总览 ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = '📊 总览'
    ws1.sheet_view.showGridLines = False
    ws1.row_dimensions[1].height = 8

    # 大标题
    ws1.merge_cells('A2:K2')
    t = ws1['A2']
    t.value = '📈 本地推投流月报  2026年6月  (06-01 ~ 06-28)'
    t.font = Font(name='微软雅黑', bold=True, size=16, color=DARK_BLUE[2:])
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[2].height = 40

    ws1.merge_cells('A3:K3')
    sub = ws1['A3']
    sub.value = f'数据截止：2026-06-28 | 覆盖账户：25个活跃账户（3个零消耗账户未计入均值）| 生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}'
    sub.font = Font(name='微软雅黑', size=9, color='888888')
    sub.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[3].height = 20

    # KPI 卡片区
    ws1.row_dimensions[5].height = 28
    ws1.row_dimensions[6].height = 40
    ws1.row_dimensions[7].height = 26

    kpi_headers = ['总消耗', '日均消耗', '总展示', '总点击', 'CTR', '私信咨询', '私信留资', 'CPL(留资)', '活跃账户', 'CPM', 'CPC']
    kpi_cells = ['B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'J5', 'K5', 'L5']

    c = conn.cursor()
    c.execute('''
        SELECT ROUND(SUM(stat_cost),2), SUM(show_cnt), SUM(click_cnt),
               ROUND(SUM(click_cnt)*1.0/NULLIF(SUM(show_cnt),0)*100,4),
               SUM(message_action_cnt), SUM(clue_message_count),
               COUNT(DISTINCT account_id)
        FROM account_reports
        WHERE stat_date>=? AND stat_date<=? AND delivery_type='total'
    ''', (START_DATE, END_DATE))
    row = c.fetchone()
    total_cost, shows, clicks, ctr, msg_action, clue_msg, acc_cnt = row
    cpl = round(total_cost / clue_msg, 2) if clue_msg else 0
    cpm = round(total_cost / shows * 1000, 2) if shows else 0
    cpc = round(total_cost / clicks, 2) if clicks else 0
    daily_avg = round(total_cost / 28, 2)

    kpi_values = [
        f'¥{total_cost:,.0f}', f'¥{daily_avg:,.0f}', f'{shows:,}', f'{clicks:,}',
        f'{ctr:.2f}%', f'{msg_action:,}', f'{clue_msg:,}',
        f'¥{cpl:,.0f}', str(acc_cnt), f'¥{cpm:,.2f}', f'¥{cpc:,.2f}'
    ]

    kpi_bg = [MID_BLUE, LIGHT_BLUE, LIGHT_BLUE, LIGHT_BLUE, LIGHT_BLUE,
              'FFe8f5e9', 'FFe8f5e9', YELLOW, GREY_BG, GREY_BG, GREY_BG]

    for i, (h, v, bg) in enumerate(zip(kpi_headers, kpi_values, kpi_bg)):
        col = i + 2
        hcell = ws1.cell(row=5, column=col, value=h)
        style_header(hcell, bg='FF4a6fa5' if i == 0 else 'FF6c8ebf', fg=WHITE, size=9)
        vcell = ws1.cell(row=6, column=col, value=v)
        vcell.font = Font(name='微软雅黑', bold=True, size=13 if i == 0 else 11,
                          color=DARK_BLUE[2:] if i == 0 else '333333')
        vcell.fill = PatternFill('solid', fgColor=bg[2:])
        vcell.alignment = Alignment(horizontal='center', vertical='center')
        vcell.border = thin_border()
        hcell.border = thin_border()
        ws1.column_dimensions[get_column_letter(col)].width = 14

    ws1['A5'].value = ''
    ws1['A6'].value = ''
    ws1.column_dimensions['A'].width = 2

    # ─── Sheet 2: 每日趋势 ────────────────────────────────────────────────
    ws2 = wb.create_sheet('📅 每日趋势')
    ws2.sheet_view.showGridLines = False
    ws2.row_dimensions[1].height = 8

    ws2.merge_cells('A2:J2')
    t2 = ws2['A2']
    t2.value = '每日消耗 & 留资趋势（2026-06-01 ~ 06-28）'
    t2.font = Font(name='微软雅黑', bold=True, size=14, color=DARK_BLUE[2:])
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[2].height = 36

    headers2 = ['日期', '消耗(¥)', '展示', '点击', 'CTR(%)', '咨询', '留资', 'CPM(¥)', 'CPL(¥)', '活跃账户']
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=ci, value=h)
        style_header(cell, bg=DARK_BLUE)
        ws2.row_dimensions[4].height = 24

    c.execute('''
        SELECT stat_date,
               ROUND(SUM(stat_cost),2), SUM(show_cnt), SUM(click_cnt),
               ROUND(SUM(click_cnt)*1.0/NULLIF(SUM(show_cnt),0)*100,4),
               SUM(message_action_cnt), SUM(clue_message_count),
               COUNT(DISTINCT account_id)
        FROM account_reports
        WHERE stat_date>=? AND stat_date<=? AND delivery_type='total'
        GROUP BY stat_date ORDER BY stat_date
    ''', (START_DATE, END_DATE))
    daily_rows = c.fetchall()

    daily_costs = [r[1] for r in daily_rows]
    avg_cost = sum(daily_costs) / len(daily_costs)
    max_cost = max(daily_costs)

    for ri, r in enumerate(daily_rows, 5):
        cost, show, click, ctr_d, msg, clue, acc = r[1], r[2], r[3], r[4], r[5], r[6], r[7]
        cpm_d = round(cost / show * 1000, 2) if show else 0
        cpl_d = round(cost / clue, 2) if clue else 0
        row_vals = [r[0], cost, show, click, ctr_d, msg, clue, cpm_d, cpl_d, acc]
        bg = 'FFFFFFFF' if ri % 2 == 0 else GREY_BG
        for ci, v in enumerate(row_vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            style_cell(cell, bg=bg, center=(ci > 1))
            cell.border = thin_border()
            if ci == 2 and cost >= max_cost * 0.95:
                cell.fill = PatternFill('solid', fgColor='FFd4edda')
                cell.font = Font(name='微软雅黑', bold=True, size=10, color=GREEN[2:])
            elif ci == 2 and cost < avg_cost * 0.9:
                cell.fill = PatternFill('solid', fgColor='FFfff3cd')
        ws2.row_dimensions[ri].height = 20

    set_col_widths(ws2, [14, 14, 14, 10, 10, 10, 10, 12, 12, 12])
    freeze(ws2, 'B5')

    # ─── Sheet 3: 账户排行 ────────────────────────────────────────────────
    ws3 = wb.create_sheet('🏆 账户排行')
    ws3.sheet_view.showGridLines = False
    ws3.row_dimensions[1].height = 8

    ws3.merge_cells('A2:L2')
    t3 = ws3['A2']
    t3.value = '账户投放综合排行（2026-06-01 ~ 06-28）'
    t3.font = Font(name='微软雅黑', bold=True, size=14, color=DARK_BLUE[2:])
    t3.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[2].height = 36

    headers3 = ['排名', '账户名称', '活跃天', '消耗(¥)', '占比%', '展示', '点击', 'CTR%', '咨询', '留资', 'CPL(¥)', '消耗/天']
    for ci, h in enumerate(headers3, 1):
        cell = ws3.cell(row=4, column=ci, value=h)
        style_header(cell, bg=DARK_BLUE)
    ws3.row_dimensions[4].height = 24

    c.execute('''
        SELECT a.name,
               COUNT(DISTINCT r.stat_date) as days,
               ROUND(SUM(r.stat_cost),2),
               SUM(r.show_cnt), SUM(r.click_cnt),
               ROUND(SUM(r.click_cnt)*1.0/NULLIF(SUM(r.show_cnt),0)*100,4),
               SUM(r.message_action_cnt), SUM(r.clue_message_count)
        FROM account_reports r
        JOIN accounts a ON r.account_id=a.account_id
        WHERE r.stat_date>=? AND r.stat_date<=? AND r.delivery_type='total'
        GROUP BY r.account_id, a.name
        ORDER BY 3 DESC
    ''', (START_DATE, END_DATE))
    acc_rows = c.fetchall()

    for ri, r in enumerate(acc_rows, 5):
        rank = ri - 4
        name, days, cost, show, click, ctr_a, msg, clue = r
        pct = round(cost / total_cost * 100, 2) if total_cost else 0
        cpl_a = round(cost / clue, 0) if clue else 0
        daily_c = round(cost / days, 0) if days else 0
        bg = 'FFFFFFFF' if rank % 2 == 0 else GREY_BG
        medal_bg = {1: 'FFffd700', 2: 'FFc0c0c0', 3: 'FFcd7f32'}

        vals = [rank, name, days, cost, pct, show, click, ctr_a, msg, clue, cpl_a, daily_c]
        for ci, v in enumerate(vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=v)
            if ci == 1 and rank <= 3:
                style_cell(cell, bold=True, center=True,
                           bg=medal_bg.get(rank, GREY_BG))
            else:
                style_cell(cell, bg=bg, center=(ci != 2))
            cell.border = thin_border()
            # format cost columns
            if ci in (4, 11, 12):
                cell.number_format = '#,##0.00'
            if ci == 5:
                cell.number_format = '0.00"%"'
        ws3.row_dimensions[ri].height = 20

    set_col_widths(ws3, [6, 24, 8, 14, 8, 14, 10, 8, 8, 8, 10, 12])
    freeze(ws3, 'C5')

    # ─── Sheet 4: 素材 TOP50 ─────────────────────────────────────────────
    ws4 = wb.create_sheet('🎯 素材TOP50')
    ws4.sheet_view.showGridLines = False
    ws4.row_dimensions[1].height = 8

    ws4.merge_cells('A2:L2')
    t4 = ws4['A2']
    t4.value = '素材投放效果 TOP50（2026-06-01 ~ 06-28，按消耗降序）'
    t4.font = Font(name='微软雅黑', bold=True, size=14, color=DARK_BLUE[2:])
    t4.alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[2].height = 36

    headers4 = ['排名', '素材ID', '所属账户', '活跃天', '消耗(¥)', '展示', '点击', 'CTR%', '咨询', '留资', 'CPL(¥)', '决策']
    for ci, h in enumerate(headers4, 1):
        cell = ws4.cell(row=4, column=ci, value=h)
        style_header(cell, bg=DARK_BLUE)
    ws4.row_dimensions[4].height = 24

    c.execute('''
        SELECT m.material_id, a.name,
               SUM(m.stat_cost) as cost,
               SUM(m.show_cnt), SUM(m.click_cnt),
               ROUND(SUM(m.click_cnt)*1.0/NULLIF(SUM(m.show_cnt),0)*100,4),
               SUM(m.message_action_cnt), SUM(m.clue_message_count),
               ROUND(SUM(m.stat_cost)/NULLIF(SUM(m.clue_message_count),0),2),
               COUNT(DISTINCT m.stat_date)
        FROM material_reports m
        JOIN accounts a ON m.account_id=a.account_id
        WHERE m.stat_date>=? AND m.stat_date<=? AND m.stat_cost>0
        GROUP BY m.material_id
        ORDER BY cost DESC
        LIMIT 50
    ''', (START_DATE, END_DATE))
    mat_rows = c.fetchall()

    # Calculate median CPL for decision scoring
    all_cpls = [r[8] for r in mat_rows if r[8] and r[8] > 0]
    med_cpl = sorted(all_cpls)[len(all_cpls)//2] if all_cpls else 50

    def decision_tag(cost, ctr, clue, cpl):
        if not cpl: return ('观察', 'FFFFE0B2')
        if cpl < med_cpl * 0.8 and ctr and ctr > 0.8:
            return ('⭐ 放量', 'FFe8f5e9')
        elif cpl < med_cpl * 1.0:
            return ('✅ 潜力', 'FFe3f2fd')
        elif cpl < med_cpl * 1.3:
            return ('👀 观察', 'FFFFFF9e')
        else:
            return ('❌ 淘汰', 'FFffebee')

    for ri, r in enumerate(mat_rows, 5):
        rank = ri - 4
        mid, acc_name, cost, show, click, ctr_m, msg, clue, cpl_m, days = r
        tag, tag_bg = decision_tag(cost, ctr_m, clue, cpl_m)
        bg = 'FFFFFFFF' if rank % 2 == 0 else GREY_BG

        vals = [rank, mid, acc_name, days, round(cost, 2), show, click, ctr_m, msg, clue, cpl_m, tag]
        for ci, v in enumerate(vals, 1):
            cell = ws4.cell(row=ri, column=ci, value=v)
            if ci == 12:
                style_cell(cell, center=True, bg=tag_bg, bold=True)
            else:
                style_cell(cell, bg=bg, center=(ci not in (2, 3)))
            cell.border = thin_border()
        ws4.row_dimensions[ri].height = 20

    set_col_widths(ws4, [6, 20, 22, 8, 14, 14, 10, 8, 8, 8, 10, 10])
    freeze(ws4, 'C5')

    # ─── Sheet 5: 投放单元 TOP30 ──────────────────────────────────────────
    ws5 = wb.create_sheet('📦 单元TOP30')
    ws5.sheet_view.showGridLines = False
    ws5.row_dimensions[1].height = 8

    ws5.merge_cells('A2:K2')
    t5 = ws5['A2']
    t5.value = '投放单元效果 TOP30（2026-06-01 ~ 06-28，按消耗降序）'
    t5.font = Font(name='微软雅黑', bold=True, size=14, color=DARK_BLUE[2:])
    t5.alignment = Alignment(horizontal='center', vertical='center')
    ws5.row_dimensions[2].height = 36

    headers5 = ['排名', '单元名称', '所属账户', '项目', '活跃天', '消耗(¥)', '展示', '点击', 'CTR%', '留资', 'CPL(¥)']
    for ci, h in enumerate(headers5, 1):
        cell = ws5.cell(row=4, column=ci, value=h)
        style_header(cell, bg=DARK_BLUE)
    ws5.row_dimensions[4].height = 24

    c.execute('''
        SELECT p.promotion_name, a.name, p.project_name,
               COUNT(DISTINCT p.stat_date) as days,
               ROUND(SUM(p.stat_cost),2),
               SUM(p.show_cnt), SUM(p.click_cnt),
               ROUND(SUM(p.click_cnt)*1.0/NULLIF(SUM(p.show_cnt),0)*100,4),
               SUM(p.clue_message_count),
               ROUND(SUM(p.stat_cost)/NULLIF(SUM(p.clue_message_count),0),2)
        FROM promotion_reports p
        JOIN accounts a ON p.account_id=a.account_id
        WHERE p.stat_date>=? AND p.stat_date<=?
        GROUP BY p.promotion_id
        ORDER BY 5 DESC
        LIMIT 30
    ''', (START_DATE, END_DATE))
    promo_rows = c.fetchall()

    for ri, r in enumerate(promo_rows, 5):
        rank = ri - 4
        pname, acc_name, proj_name, days, cost, show, click, ctr_p, clue, cpl_p = r
        bg = 'FFFFFFFF' if rank % 2 == 0 else GREY_BG
        vals = [rank, pname or '-', acc_name, proj_name or '-', days, cost, show, click, ctr_p, clue, cpl_p]
        for ci, v in enumerate(vals, 1):
            cell = ws5.cell(row=ri, column=ci, value=v)
            style_cell(cell, bg=bg, center=(ci not in (2, 3, 4)))
            cell.border = thin_border()
        ws5.row_dimensions[ri].height = 20

    set_col_widths(ws5, [6, 28, 22, 20, 8, 12, 14, 10, 8, 8, 10])
    freeze(ws5, 'C5')

    # ─── Sheet 6: CPL 效率分析 ────────────────────────────────────────────
    ws6 = wb.create_sheet('💡 账户效率')
    ws6.sheet_view.showGridLines = False
    ws6.row_dimensions[1].height = 8

    ws6.merge_cells('A2:J2')
    t6 = ws6['A2']
    t6.value = 'CPL 效率分析 & 账户建议（2026-06-01 ~ 06-28）'
    t6.font = Font(name='微软雅黑', bold=True, size=14, color=DARK_BLUE[2:])
    t6.alignment = Alignment(horizontal='center', vertical='center')
    ws6.row_dimensions[2].height = 36

    headers6 = ['账户', '消耗', '留资', 'CPL', '咨询', '留资率%', '消耗占比%', '日均消耗', 'CTR%', '评级']
    for ci, h in enumerate(headers6, 1):
        cell = ws6.cell(row=4, column=ci, value=h)
        style_header(cell, bg=DARK_BLUE)
    ws6.row_dimensions[4].height = 24

    c.execute('''
        SELECT a.name,
               ROUND(SUM(r.stat_cost),2),
               SUM(r.clue_message_count),
               ROUND(SUM(r.stat_cost)/NULLIF(SUM(r.clue_message_count),0),2),
               SUM(r.message_action_cnt),
               ROUND(SUM(r.clue_message_count)*1.0/NULLIF(SUM(r.message_action_cnt),0)*100,2),
               ROUND(SUM(r.click_cnt)*1.0/NULLIF(SUM(r.show_cnt),0)*100,4),
               COUNT(DISTINCT r.stat_date)
        FROM account_reports r
        JOIN accounts a ON r.account_id=a.account_id
        WHERE r.stat_date>=? AND r.stat_date<=? AND r.delivery_type='total' AND r.stat_cost>0
        GROUP BY r.account_id, a.name
        ORDER BY 4 ASC
    ''', (START_DATE, END_DATE))
    eff_rows = c.fetchall()

    cpls = [r[3] for r in eff_rows if r[3] and r[3] > 0]
    avg_cpl_all = sum(cpls) / len(cpls) if cpls else 50

    def grade(cpl):
        if not cpl: return ('N/A', GREY_BG)
        if cpl < avg_cpl_all * 0.7: return ('⭐⭐⭐ 优秀', 'FFe8f5e9')
        elif cpl < avg_cpl_all * 0.9: return ('⭐⭐ 良好', 'FFe3f2fd')
        elif cpl < avg_cpl_all * 1.1: return ('⭐ 正常', YELLOW)
        elif cpl < avg_cpl_all * 1.3: return ('⚠️ 偏高', 'FFFFE0B2')
        else: return ('🔴 超高', 'FFffebee')

    for ri, r in enumerate(eff_rows, 5):
        name, cost, clue, cpl, msg, conv_rate, ctr_e, days = r
        pct = round(cost / total_cost * 100, 2) if total_cost else 0
        daily_c = round(cost / days, 0) if days else 0
        g_label, g_bg = grade(cpl)
        bg = 'FFFFFFFF' if (ri-5) % 2 == 0 else GREY_BG

        vals = [name, cost, clue, cpl, msg, conv_rate, pct, daily_c, ctr_e, g_label]
        for ci, v in enumerate(vals, 1):
            cell = ws6.cell(row=ri, column=ci, value=v)
            if ci == 10:
                style_cell(cell, center=True, bg=g_bg, bold=True)
            else:
                style_cell(cell, bg=bg, center=(ci != 1))
            cell.border = thin_border()
        ws6.row_dimensions[ri].height = 20

    # 底部说明
    note_row = len(eff_rows) + 6
    ws6.merge_cells(f'A{note_row}:J{note_row}')
    note = ws6.cell(row=note_row, column=1)
    note.value = f'※ 均值CPL = ¥{avg_cpl_all:.0f} | 评级标准：优秀<{avg_cpl_all*0.7:.0f} / 良好<{avg_cpl_all*0.9:.0f} / 正常<{avg_cpl_all*1.1:.0f} / 偏高<{avg_cpl_all*1.3:.0f} / 超高≥{avg_cpl_all*1.3:.0f}'
    note.font = Font(name='微软雅黑', size=9, color='666666', italic=True)
    note.alignment = Alignment(horizontal='left', vertical='center')
    ws6.row_dimensions[note_row].height = 20

    set_col_widths(ws6, [24, 14, 8, 10, 8, 10, 10, 12, 8, 14])
    freeze(ws6, 'B5')

    # 确保 outputs 目录存在
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, '6月投流月报_2026-06-01_06-28.xlsx')
    wb.save(out_path)
    return out_path


if __name__ == '__main__':
    conn = sqlite3.connect(DB_PATH)
    path = build(conn)
    conn.close()
    print(f'✅ 月报已生成: {path}')
